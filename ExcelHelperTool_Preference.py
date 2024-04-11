from openpyxl import load_workbook
import os, math

# TODO: Remove useless code?

# ============= Workshop variables to be customized =============================== #
excel_file_name = 'WT_Einteilung_ANO_.xlsx' #'StudentWorkshop_SampleExcelSheet.xlsx'
sheetname = 'WerkstattStundenplan'

last_name_column = 1 #A
first_name_column = 2 #B
name_start_row = 2 # the row where student's name start
workshop_name_start_column = 4 #D 
student_schedule_column = 15 #O

num_workshops_for_students = 5  # Number of workshops each student must take
num_workshop_rounds = 5   # Number of sessions/rounds each workshop provides

# ============== Do not edit the variables beyond this point! ===================== #
current_directory = os.path.dirname(os.path.abspath(__file__))
file_path = os.path.join(current_directory, excel_file_name)
workbook = load_workbook(file_path)
sheet = workbook.get_sheet_by_name(sheetname) 
debug_mode = True

# Load workshop name in order from excel
def load_workshop_names_from_excel():
    list = []
    current_column = workshop_name_start_column

    while True:
        workshop_name = sheet.cell(row = 1, column = current_column).value
        if workshop_name is not None:
            list.append(workshop_name)
            current_column += 1
        else:
            pref_end_column = current_column - 1
            break
    return list, pref_end_column

# Load student and group name from excel
def load_student_names_from_excel():
    names = []
    current_row = name_start_row
    
    while True:
        last_name = sheet.cell(row=current_row, column=last_name_column).value
        first_name = sheet.cell(row=current_row, column=first_name_column).value
        
        if last_name is not None and first_name is not None:
            full_name = f"{last_name} {first_name}"
            names.append(full_name)
            current_row += 1
        else:
            break
    return names

# Static workshop_nameiables
student_names = load_student_names_from_excel()
workshop_list, preference_end_column = load_workshop_names_from_excel()

# DEBUG: Make this flexible so it clears until the next blank row & column
# Erase the specified range of cells for student schedules
def Cells_Cleanup(start_row, start_col, end_row = None, end_col = None):
    #if end row or end col is not definied, the clean up will be executed until a blank row/col is met.
    if not end_row: 
        end_row = start_row + len(student_names)
    if not end_col:
        end_col = None

    for row in sheet.iter_rows(min_row=end_row + 1, min_col=4, max_col=4):
        if row[0].value is not None:
            end_row +=1
        else:
            break

    # Clean up cells from start_col to end_col until end_row
    for row in sheet.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
        for cell in row:
            cell.value = None

    workbook.save(excel_file_name)

# Create a dictionary with student name & their preference
# Also produce a list of student with no preference
def Fetch_Student_Preference_List():
    dict_student_pref = {}
    students_with_no_pref = []
    current_row = name_start_row
    workshop_choices = preference_end_column - workshop_name_start_column + 1 #calculate how many workshops to choose from

    # Loop through all students until the next empty row
    while True:
        last_name = sheet.cell(row=current_row, column=last_name_column).value
        first_name = sheet.cell(row=current_row, column=first_name_column).value
        
        # for each student:
        if last_name is not None and first_name is not None:
            full_name = f"{last_name} {first_name}"
            curr_pref_column = workshop_name_start_column
            int_preference_list = []

            # loop through all preference column and construct a list 
            for i in range(workshop_choices):
                current_pref = sheet.cell(row = current_row, column = curr_pref_column).value
                int_preference_list.append(current_pref)

                curr_pref_column += 1

                # NOTE: make sure to take into consideration some kids doesn't fill everything
                # CURRENTLY THE CODE WILL SKIP STUDENT THAT DOES NOT HAVE PREF FOR ALL WORKSHOP
                if current_pref is None:
                    int_preference_list = []
                    break
            
            # sort student and pref pair to dict ; student with no pref to list
            if not int_preference_list:
                students_with_no_pref.append(full_name)
            else:
                preference_list = Sort_Preference_List(int_preference_list)
                dict_student_pref[full_name] = preference_list

            current_row += 1
        else:
            break
    
    return dict_student_pref, students_with_no_pref

def sublist_length(sublist):
    return len(sublist) if sublist is not None else float('inf')  # Use infinity for None to ensure they are placed at the end

# Translate the value to workshop name and arrange them from favorite -> least favorite
def Sort_Preference_List(int_preflist):
    # int_preflist will turn from [1,3,6,10,...] to [1,2,3,4,5,...]
    temp_workshop_list = workshop_list
    zipped_lists = zip(int_preflist, temp_workshop_list)
    sorted_zipped_lists = sorted(zipped_lists, key=lambda x: x[0])
    sorted_workshop_preference = [item[1] for item in sorted_zipped_lists]

    # translate the number to workshop name [IT, Gastro Textil]
    return sorted_workshop_preference
student_preference_dict, students_with_no_preference = Fetch_Student_Preference_List()

# Calculate how many student should be in each workshop, if this num exceed 13 then keep it 13.
def Calculate_Max_Students_Per_Workshop():
    number_of_students = len(student_names)
    number_of_workshops = len(workshop_list)
    maximum = (number_of_students * num_workshops_for_students) / (number_of_workshops * num_workshop_rounds)
    maximum = math.ceil(maximum)
    #if maximum > 13: maximum = 13
    maximum -= 1
    return maximum

dict_student_schedules = {key: {} for key in student_names}  # Keep track of each student and their workshop
dict_workshop_students = {} # Consist workshop name and 5 lists of student in each session: {Textil:[[student A,B,C],[student D,E,F],[student G,H,I]]; }
max_workshop_size = Calculate_Max_Students_Per_Workshop()
print(max_workshop_size)

# Return bool to check if all sessions of a workshop is full
def Remove_Full_Workshop(max_per_workshop, available_workshops):
    maximum_total_students = max_per_workshop * num_workshop_rounds
    for key,list_of_sublists in dict_workshop_students.items():
        total_students = sum(len(sublist) for sublist in list_of_sublists)
        if total_students >= maximum_total_students:
            if key in available_workshops:
                available_workshops.remove(key)
                print(key,"is full!")

def Get_Student_Missing_Session(list_to_compare):
    full_schedule_integer = set(range(5)) # a set of 0-4 to check which session  missing in the list
    missing_session = full_schedule_integer - set(list_to_compare)
    missing_session = missing_session.pop()
    return missing_session

# TODO: the full class should not appear to be exchanged if the current session is full.

# go through EACH in workshops already taken and see if any can be replaced with another session
def Rearrange_Session_For_Repeated_Class(student, workshops_already_taken, missing_session):
    for workshop, index in workshops_already_taken.items():
        # make a dict of workshop/sessions with the session index
        if len(dict_workshop_students[workshop][missing_session]) > max_workshop_size:
            continue
        
        new_session_dict = {} # {workshopA : 11(number of students), ...}
        for key, list_of_lists in dict_workshop_students.items():
            if key not in workshops_already_taken:    # make sure not repeat the workshops
                size = len(list_of_lists[index])
                new_session_dict[key] = size
        
        # assign the index with the smallest workshop, if it's not full. Otherwise continue the loop
        smallest_workshop = min(new_session_dict.items(), key=lambda x: x[1]) # return the pair ('workshopA', 11)

        if smallest_workshop[1] < max_workshop_size:
            # repeat the workshop for the missing session, and take a different workshop for the repeated session
            dict_workshop_students[smallest_workshop[0]][index].append(student)   # exchanged workshop
            dict_workshop_students[workshop][missing_session].append(student)   # repeat workshop

            # remove student from the session where workshop is repeated
            dict_workshop_students[workshop][index].remove(student)
            # remove the student's schedule on the index
            dict_student_schedules[student] = {key: value for key, value in dict_student_schedules[student].items() if value != index}

            # update the workshops in the student's timetable
            dict_student_schedules[student].update({smallest_workshop[0] : index})
            dict_student_schedules[student].update({workshop : missing_session})
            break
        # if the workshop is full, move on to the next repeated workshop and try again
    print(student,"took",smallest_workshop[0],"and rearranged",workshop)
    
# Main logic to sort students to workshop based on their preference
def Sort_Student_to_Workshop_by_Preference():
    global max_workshop_size

    for a in range(len(workshop_list)):
        w_name = workshop_list[a]
        w_list = []
        for b in range(num_workshop_rounds):
            w_list.append([])
        dict_workshop_students[w_name] = w_list

    #keep track of the workshops still available
    available_workshops = workshop_list.copy()
    
    # Assign workshops according to student's preference
    for i in range(num_workshop_rounds):
        # loop through all students
        for student in student_preference_dict:
            workshop_reorganized = False

            # Get their 1st choice
            student_choice = student_preference_dict[student][0]

            # retrieve the list of sessions from the chosen workshop
            chosen_workshop_sessions = dict_workshop_students[student_choice]
            
            # Filter out the index of session that student's are already occupied with
            workshop_indexes = list(dict_student_schedules[student].values())
            workshop_indexes.sort()

            available_sessions_to_take = chosen_workshop_sessions.copy()
            for i in workshop_indexes:
                available_sessions_to_take[i] = None

            # Find the length of the smallest workshop in the available ones, sort student to that workshop
            session_with_least_students = min(filter(lambda x: x is not None, available_sessions_to_take), key=sublist_length)

            # keep looping through the list of student's pref list until they're filled to the next workshop
            while len(session_with_least_students) >= max_workshop_size:
                # remove the preference for occupied workshop
                del student_preference_dict[student][0]

                if not student_preference_dict[student]:
                    workshop_reorganized = True
                    missing_session = Get_Student_Missing_Session(workshop_indexes)
                    Rearrange_Session_For_Repeated_Class(student, dict_student_schedules[student], missing_session)
                    break
                #get the next available workshop on the pref list
                student_choice = student_preference_dict[student][0]
                
                chosen_workshop_sessions = dict_workshop_students[student_choice]
                available_sessions_to_take = chosen_workshop_sessions.copy()
                for index in workshop_indexes:
                    available_sessions_to_take[index] = None

                session_with_least_students = min(filter(lambda x: x is not None, available_sessions_to_take), key=sublist_length)
            
            if workshop_reorganized: continue

            # remove the workshop from their list
            del student_preference_dict[student][0]

            # assign student to that session
            workshop_index = available_sessions_to_take.index(session_with_least_students)
            dict_workshop_students[student_choice][workshop_index].append(student)
            
            # put the workshop in the student's timetable
            dict_student_schedules[student].update({student_choice : workshop_index})
            # go to the next student

    # Remove_Full_Workshop(max_workshop_size,dict_workshop_students,available_workshops)
    max_workshop_size += 1 # real maximum
    
    # Assign workshop for students with no preference
    print("student with no pref:",students_with_no_preference)
    #'''
    for i in range(num_workshop_rounds):
        # loop through the available workshop and get rid of the actual full one here
        Remove_Full_Workshop(max_workshop_size,available_workshops)

        for student in students_with_no_preference:
            workshop_reorganized = False
            # get the list of workshop taken by current student, remove them from the available choice 
            workshops_already_taken = dict_student_schedules[student]
            workshops_to_choose = [x for x in available_workshops if x not in workshops_already_taken]
            
            # TODO: Case where student can only repeat class and the only available ones are all repeated. 
                # (meaning they can't switch their existing class with the available one.)
            if not workshops_to_choose:
                print("no more workshop to choose for",student)
                print("available workshops:",available_workshops,"workshops already taken:",workshops_already_taken)
                continue

            # Create a temporary copy of the dict, remove the workshop that's already taken by student.
            new_dict_workshop_students = {key: value for key, value in dict_workshop_students.items() if key in workshops_to_choose}
            
            # Calculate the sum of total students for each workshop in the new dict
            total_student_in_each_workshop = {key: sum(len(sublist) for sublist in value) for key, value in new_dict_workshop_students.items()}

            # Pick the workshop with the least number of students in total
            workshop_with_least_students = min(total_student_in_each_workshop, key=total_student_in_each_workshop.get)
            chosen_workshop_sessions = dict_workshop_students[workshop_with_least_students]

            # access the index of workshops students are taking
            workshop_indexes = list(dict_student_schedules[student].values())
            # remove those from consideration in this min()???
            workshop_indexes.sort()
            available_sessions_to_take = chosen_workshop_sessions.copy()

            for index in workshop_indexes:
                available_sessions_to_take[index] = None

            # take the session with least number of people
            session_with_least_students = min(filter(lambda x: x is not None, available_sessions_to_take), key=sublist_length)
            
            while len(session_with_least_students) >= max_workshop_size:
                total_student_in_each_workshop.pop(workshop_with_least_students) # dict {workshop name : number of total students}
                
                if not total_student_in_each_workshop:
                    # get the index of the session student need to take
                    workshop_reorganized = True

                    # go through EACH in workshops already taken and see if any can be replaced with another session
                    missing_session = Get_Student_Missing_Session(workshop_indexes)
                    Rearrange_Session_For_Repeated_Class(student, workshops_already_taken, missing_session)
                    break
                
                # Pick the workshop with the least number of students in total
                workshop_with_least_students = min(total_student_in_each_workshop, key=total_student_in_each_workshop.get)
                chosen_workshop_sessions = dict_workshop_students[workshop_with_least_students]

                available_sessions_to_take = chosen_workshop_sessions.copy()

                for index in workshop_indexes:  #index is already previously calculated
                    available_sessions_to_take[index] = None
                
                session_with_least_students = min(filter(lambda x: x is not None, available_sessions_to_take), key=sublist_length)
            
            if workshop_reorganized: continue

            # Assign student to the session
            workshop_index = available_sessions_to_take.index(session_with_least_students)
            dict_workshop_students[workshop_with_least_students][workshop_index].append(student)
            
            # put the workshop in the student's timetable
            dict_student_schedules[student].update({workshop_with_least_students : workshop_index})
    #'''
    # Print the summary
    if debug_mode:
        print("**********WORKSHOP SUMMARY*****************")
        for key, value in dict_workshop_students.items():
            print(key)
            for sublist in value:
                print(len(sublist),":",sublist)
        
        
        print("\n\n**********STUDENT SCHEDULE*****************")
        for key, value in dict_student_schedules.items():
            print(key)
            print(len(value),":",value)
    
    return dict_student_schedules  
    
# Rearrange student's schedule based on the order of the workshop
def Rearrange_Student_Schedule(student_schedule_dict):
    # student_preference_dict: student_A:{[]:2,[],10}
    converted_dict = {}
    for key, value in student_schedule_dict.items():
        sorted_subdict = sorted(value.items(), key=lambda x: x[1])
        converted_list = [item[0] for item in sorted_subdict]
        converted_dict[key] = converted_list

    return converted_dict
    
# Update excel with list of workshops for each student 
def Excel_Update_Student_Schedule(student_schedule_dict):
    row = name_start_row # starting row 

    for student, workshops in student_schedule_dict.items():
        col = student_schedule_column # Write each element from the list to columns D to H

        for w in workshops:
            sheet.cell(row=row, column=col, value = w)
            col += 1
        
        row += 1  # Move to the next row for the next list
    workbook.save(file_path)

# Execution code
student_schedule_dict = Sort_Student_to_Workshop_by_Preference()
student_schedule = Rearrange_Student_Schedule(student_schedule_dict)

Cells_Cleanup(2, 15, None, 19)
Excel_Update_Student_Schedule(student_schedule)